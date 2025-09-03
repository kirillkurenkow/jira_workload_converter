import os.path
from argparse import ArgumentParser, ArgumentTypeError
import calendar
import logging
import re
from datetime import (
    date,
    datetime,
    timedelta,
)
from typing import (
    Any,
    Dict,
    List,
    Literal,
    Optional,
    Tuple,
)

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Types
_T_GENERATED_DATA = List[List[Any]]
_T_MERGE_DATA = List[Tuple[int, int, int, int]]
_T_CORDS = Tuple[int, int]
_T_STYPE_BORDER = List[Dict[Tuple[_T_CORDS, _T_CORDS] | Literal['full'], Border | bool]]
_T_STYLE_DATA = Dict[Literal['all', 'borders'] | Tuple[_T_CORDS, _T_CORDS], Dict[str, Any] | _T_STYPE_BORDER]


def get_weeks_for_year(year: Optional[int] = None) -> Dict[date, date]:
    """
    Get weeks dict for year

    :param year: Year

    :return: Weeks dict for year:
    {
        week_1_start_dt: week_1_end_dt,
        ...
    }
    """
    if year is None:
        year = datetime.today().year

    calendar_object = calendar.Calendar(0)
    weeks = [calendar_object.monthdatescalendar(year, month_i) for month_i in range(1, 13)]
    weeks = [(x[0], x[-1]) for row in weeks for x in row]
    weeks = {x[0]: x[1] for x in (sorted(set(weeks)))}

    return weeks


class Const:
    class Scheduler:
        WORK_HOURS_PER_DAY = 8
        WORK_DAYS_PER_WEEK = 5


class Task:
    def __init__(
        self,
        name: str,
        key: str,
        date_start: date,
        date_end: date,
        priority: Optional[str] = None,
        status: Optional[str] = None,
        estimate: Optional[int] = None,
    ):
        self._name = name
        self._key = key
        self._priority = priority
        self._status = status
        self._estimate = estimate
        self._date_start = date_start
        self._date_end = date_end

        LOGGER.info(f'Task created: {dict(self)}')

    @property
    def name(self) -> str:
        return self._name

    @property
    def key(self) -> str:
        return self._key

    @property
    def date_start(self) -> date:
        return self._date_start

    @property
    def date_end(self) -> date:
        return self._date_end

    @property
    def priority(self) -> Optional[str]:
        return self._priority

    @property
    def status(self) -> Optional[str]:
        return self._status

    @property
    def estimate(self) -> Optional[int]:
        return self._estimate

    @staticmethod
    def is_jira_key(key: str) -> bool:
        """
        Check if string is actual JIRA task key

        :param str key: Key to check

        :return: String is actual JIRA task key
        :rtype:  bool
        """
        pattern_jirauser = r'JIRAUSER\d+'
        pattern_user = r'\w\.\w+'

        result = not (re.match(pattern_user, key) or re.match(pattern_jirauser, key))

        return result

    def __iter__(self):
        result = {x: self.__getattribute__(x) for x in dir(self) if not x.startswith('_')}
        return iter(result.items())

    def __repr__(self):
        return f'Task[{self.key}]'

    def __hash__(self):
        return hash(self.key)


class Worker:
    def __init__(self, name: str, username: str, tasks: Optional[List[Task]] = None):
        self._name = name
        self._username = username
        self._tasks = [] if tasks is None else tasks.copy()

        LOGGER.info(f'Worker created: {{"name": {self._name}}}')

    @property
    def name(self) -> str:
        return self._name

    @property
    def username(self) -> str:
        return self._username

    @property
    def tasks(self) -> List[Task]:
        return self._tasks.copy()

    def add_task(self, task: Task) -> None:
        """
        Add task to worker

        :param Task task: Task object

        :return: None
        """
        self._tasks.append(task)
        LOGGER.info(f'{task} was added to {self}')

    def get_workload_by_tasks_for_year(self, year: Optional[int] = None) -> Dict[Task, List[int | float]]:
        """
        Get worker workload grouped by tasks for year

        :param year: Year

        :return: Workload by tasks for year:
        {
            task_1: [task_1_week_1_workload, ...],
            ...
        }
        """
        if year is None:
            year = datetime.today().year
        weeks = get_weeks_for_year(year)
        result = {task: [0 for _ in range(len(weeks))] for task in self._tasks}

        for task, workload in result.items():
            task_start_monday = task.date_start - timedelta(days=task.date_start.weekday())
            weeks_for_task = abs(-(task.date_end - task_start_monday).days // 7)
            hours_per_week = Const.Scheduler.WORK_DAYS_PER_WEEK * Const.Scheduler.WORK_HOURS_PER_DAY
            workload = round(float(task.estimate / (hours_per_week * weeks_for_task)), 4)
            task_start_week_index = list(weeks).index(task_start_monday)
            for i in range(weeks_for_task):
                result[task][task_start_week_index + i] = workload

        LOGGER.debug(f'{self} workload by tasks: {result}')

        return result

    def get_workload_summary_for_year(self, year: Optional[int] = None) -> List[int | float]:
        """
        Get worker workload summary for year

        :param year: Year

        :return: Workload summary for year:
        [week_1_workload_summary, ...]
        """
        if self._name == 'Тополог (не распределенный ресурс)':
            pass
        workload_by_tasks = self.get_workload_by_tasks_for_year(year)
        result = [round(sum([workload_by_tasks[task][i] for task in workload_by_tasks]), 4) for i in range(len(get_weeks_for_year(year)))]  # type: ignore  # noqa

        LOGGER.debug(f'{self}: workload summary for year: {result}')

        return result

    def __repr__(self):
        return f'Worker[{self.name}]'


class Department:
    def __init__(self, name: str, workers: Optional[List[Worker]] = None):
        self._name = name
        if isinstance(workers, list):
            self._workers = workers.copy()
        else:
            self._workers = []

    @property
    def name(self) -> str:
        return self._name

    @property
    def workers(self) -> List[Worker]:
        return self._workers.copy()

    def add_worker(self, worker: Worker) -> None:
        """
        Add worker to department workers list

        :param worker: Worker object

        :return: None
        """
        self._workers.append(worker)

    def get_department_workload(self, year: int) -> Optional[List[int | float]]:
        """
        Get department workload for year
        :param year: Year

        :return: Department workload
        :rtype:  [week_1_workload_summary, ...]
        """
        if self._workers:
            result = [worker.get_workload_summary_for_year(year) for worker in self._workers]
            result = [[x[i] for x in result] for i in range(len(result[0]))]
            result = [round(sum(x) / len(x), 4) for x in result]

            return result


class Scheduler:
    def __init__(
        self,
        departments: List[Department],
        workers: List[Worker],
        tasks: List[Task],
    ):
        self._departments = departments.copy()
        self._workers = workers.copy()
        self._tasks = tasks.copy()

    def generate_data(
        self,
        year: int = None,
    ) -> Tuple[_T_GENERATED_DATA, _T_MERGE_DATA, _T_STYLE_DATA]:
        """
        Generate data for excel

        :param year: Year

        :return: Tuple: generated data, cells merge information, cells style information
        generated data:
        [
            [row_1_col_1_value, ...],
            ...
        ]

        cells merge information:
        [
            (cell_row_i_start, cell_col_i_start, cell_row_j_end, cell_col_j_end),
            ...
        ]


        cells style information
        {
            "all": {
                '<cell_attr_name>': StyleObj,
                ...
            },
            "borders": [
                {
                    ((cell_row_i, cell_col_i), (cell_row_j, cell_col_j)): BorderObj,
                    ...
                    Optional["full"]: True,
                }
            ],
            ((cell_row_i, cell_col_i), (cell_row_j, cell_col_j)): {
                '<cell_attr_name>': StyleObj,
                ...
            },
            ...
        }
        """
        if year is None:
            year = datetime.today().year

        data = []
        weeks = get_weeks_for_year(year)
        style_data: _T_STYLE_DATA = {'borders': []}

        # Months
        months_weeks_mapping = {month: [] for month in list(calendar.month_name)}
        months_weeks_mapping.pop('')
        for week in weeks.items():
            week_start, week_end = week
            if week_start.month == week_end.month:
                months_weeks_mapping[list(months_weeks_mapping)[week_start.month - 1]].append(week)
            elif week_start.year == YEAR - 1:
                months_weeks_mapping[list(months_weeks_mapping)[0]].append(week)
            elif week_end.year == YEAR + 1:
                months_weeks_mapping[list(months_weeks_mapping)[11]].append(week)
            else:
                months_weeks_mapping[list(months_weeks_mapping)[week_start.month - 1]].append(week)
        month_headers = []
        for month_name, month_weeks in months_weeks_mapping.items():
            month_headers.append(month_name)
            month_headers.extend([''] * (len(month_weeks) - 1))

        # Headers
        base_headers = ['Worker', 'Task name', 'Task key']
        data.append(base_headers + ['Timeline'])  # Base headers

        data.append([''] * len(base_headers) + month_headers)  # Month headers
        data.append([''] * len(base_headers) + list(range(1, len(weeks) + 1)))  # Weeks headers

        # Main data
        department_cords = []
        for department in sorted(self._departments, key=lambda x: x.name):
            if department.workers and any([worker.tasks for worker in department.workers]):
                data.append([department.name] + [''] * (len(base_headers) - 1) + department.get_department_workload(year))  # noqa
                department_cords_temp = (len(data), 1)
                style_data[(department_cords_temp, department_cords_temp)] = {'font': Font(bold=True)}
                for worker in sorted(department.workers, key=lambda x: x.username):
                    data.append([worker.name] + [''] * (len(base_headers) - 1) + worker.get_workload_summary_for_year(year))  # noqa
                    for task, workload in worker.get_workload_by_tasks_for_year(year).items():
                        data.append([worker.name, task.name, task.key] + workload)  # noqa
                department_cords_temp = (department_cords_temp, (len(data), len(base_headers) + len(weeks)))
                department_cords.append(department_cords_temp)

        # Merge data
        merge_data = list()
        merge_data.append((1, 4, 1, len(base_headers) + len(weeks)))  # Timeline

        # --- Base headers
        for header_i in range(1, len(base_headers) + 1):
            merge_data.append((1, header_i, 3, header_i))

        # --- Months
        temp = 0
        for month_name, month_weeks in months_weeks_mapping.items():
            merge_index = len(base_headers) + 1 + temp
            merge_data.append((2, merge_index, 2, merge_index + len(month_weeks) - 1))
            temp += len(month_weeks)

        # --- Groups
        for row_i, row in enumerate(data[3:], 4):
            if not row[1]:
                merge_data.append((row_i, 1, row_i, 3))

        # Style data
        style_data.update({
            'all': {
                'alignment': Alignment(horizontal='center', vertical='center'),
            },  # All
            ((1, 1), (1, len(base_headers) + 1)): {
                'font': Font(bold=True, size=14),
            },  # Bold headers
            ((2, len(base_headers) + 1), (2, len(base_headers) + len(weeks) + 1)): {
                'font': Font(bold=True, size=12),
            }  # Bold months
        })

        # --- Styles by cell value
        for row_i in range(len(base_headers), len(data)):
            for col_i in range(3, len(data[row_i])):
                value = data[row_i][col_i]
                if isinstance(value, (int, float)):
                    if value > 1:
                        style = 'Bad'
                    elif 0 < value < 1:
                        style = 'Neutral'
                    else:
                        style = 'Good'
                    cords = ((row_i + 1, col_i + 1), (row_i + 1, col_i + 1))
                    style_data.setdefault(cords, {})
                    style_data[cords].update({'style': style, 'number_format': '0.00%'})
                else:
                    ...  # Log warning

        # --- Borders
        base_border_side = Side(border_style='thin', color='000000')
        bold_border_side = Side(border_style='medium', color='000000')
        base_border = Border(
            top=base_border_side,
            left=base_border_side,
            right=base_border_side,
            bottom=base_border_side,
        )
        bold_border = Border(
            top=bold_border_side,
            left=bold_border_side,
            right=bold_border_side,
            bottom=bold_border_side,
        )
        style_data['borders'].append({
            ((1, 1), (len(data), len(base_headers) + len(weeks))): base_border,
            'full': True,
        })  # Base border for all data (full)
        style_data['borders'].append({
            ((1, 1), (len(data), len(base_headers) + len(weeks))): bold_border,
        })  # Bold border for all data
        style_data['borders'].append({
            ((1, 1), (3, len(base_headers) + len(weeks))): bold_border,
        })  # Bold border for headers
        style_data['borders'].append({
            ((1, 1), (len(data), len(base_headers))): bold_border,
        })  # Bold border for base headers column
        for department_cord in department_cords:
            style_data['borders'].append({department_cord: bold_border})
        for row in data:
            LOGGER.info(f'Data generated: {row}')
        for row in merge_data:
            LOGGER.debug(f'Merge data: {row}')
        for key, value in style_data.items():
            LOGGER.debug(f'Style data: {key}: {value}')

        return data, merge_data, style_data

    @staticmethod
    def apply_border(
        worksheet: Worksheet,
        cells_range: Tuple[_T_CORDS, _T_CORDS],
        border: Border,
        full: bool = False,
    ):
        """
        Apply border by given cords

        :param worksheet: Worksheet object
        :param cells_range: Cells range
        :param border: Border object
        :param full: Apply border to every cell in given range if true or to outer ones only

        :return: None
        """
        if full:
            LOGGER.debug(f'Applying full border to {cells_range} ...')
            for row_i in range(cells_range[0][0], cells_range[1][0] + 1):
                for col_i in range(cells_range[0][1], cells_range[1][1] + 1):
                    cell = worksheet.cell(row_i, col_i)
                    cell.border = border  # noqa: It's not read-only
        else:
            LOGGER.debug(f'Applying outer border to {cells_range} ...')
            for row_i in range(cells_range[0][0], cells_range[1][0] + 1):
                cell = worksheet.cell(row_i, cells_range[0][1])
                cell.border = Border(  # noqa: It's not read-only
                    top=cell.border.top,
                    left=border.left,
                    right=cell.border.right,
                    bottom=cell.border.bottom,
                )
                cell = worksheet.cell(row_i, cells_range[1][1])
                cell.border = Border(  # noqa: It's not read-only
                    top=cell.border.top,
                    left=cell.border.left,
                    right=border.right,
                    bottom=cell.border.bottom,
                )
            for col_i in range(cells_range[0][1], cells_range[1][1] + 1):
                cell = worksheet.cell(cells_range[0][0], col_i)
                cell.border = Border(  # noqa: It's not read-only
                    top=border.top,
                    left=cell.border.left,
                    right=cell.border.right,
                    bottom=cell.border.bottom,
                )
                cell = worksheet.cell(cells_range[1][0], col_i)
                cell.border = Border(  # noqa: It's not read-only
                    top=cell.border.top,
                    left=cell.border.left,
                    right=cell.border.right,
                    bottom=border.bottom,
                )

    def write_data_to_excel(
        self,
        data: _T_GENERATED_DATA,
        merge_data: Optional[_T_MERGE_DATA] = None,
        style_data: Optional[_T_STYLE_DATA] = None,
        freeze_cell: Optional[_T_CORDS] = None,
        filename: str = "output.xlsx",
    ) -> None:
        """
        Write generated data to excel book

        :param data: Data generated by scheduler
        :param merge_data: Cells merge information
        :param style_data: Cells style information
        :param freeze_cell: Cell cords for freezing rows above and columns to the left
        :param filename: Excel book filename

        :return: None
        """
        # Create workbook
        LOGGER.info('Creating workbook ...')
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Write data
        LOGGER.info('Writing raw data ...')
        for row_i, row in enumerate(data, 1):
            for col_i, value in enumerate(row, 1):
                sheet.cell(row_i, col_i).value = value

        # Merge cells
        LOGGER.info('Merging cells ...')
        if merge_data:
            for merge_cell_data in merge_data:
                sheet.merge_cells(**dict(zip(["start_row", "start_column", "end_row", "end_column"], merge_cell_data)))

        # Add style

        # --- Apply styles by cords
        # --- --- Cords
        if style_data:
            LOGGER.info('Applying styles by given cords ...')
            for cords, style in style_data.items():
                if cords in ['all', 'borders']:
                    continue
                for row_i in range(cords[0][0], cords[1][0] + 1):  # noqa
                    for col_i in range(cords[0][1], cords[1][1] + 1):  # noqa
                        for style_obj, value in style.items():
                            sheet.cell(row_i, col_i).__setattr__(style_obj, value)

        # --- --- All
            if style_data['all']:
                LOGGER.info('Applying styles for all cells ...')
                for row_i in range(1, sheet.max_row + 1):
                    for col_i in range(1, sheet.max_column + 1):
                        for style_obj, value in style_data['all'].items():
                            sheet.cell(row_i, col_i).__setattr__(style_obj, value)

        # --- --- Borders
            if style_data['borders']:
                LOGGER.info('Applying borders by given cords ...')
                for border_style in style_data['borders']:
                    for border_cords, border in border_style.items():
                        if isinstance(border_cords, tuple):
                            self.apply_border(
                                worksheet=sheet,
                                cells_range=border_cords,
                                border=border,
                                full=border_style.get('full', False),
                            )

        # --- Adjust column widths
        LOGGER.info('Adjusting column widths ...')
        column_widths = [
            max([len(str(sheet.cell(row_index, column_index).value)) for row_index in range(4, sheet.max_row + 1)])
            for column_index in range(1, sheet.max_column + 1)
        ]
        for column_index, column_width in enumerate(column_widths, 1):
            sheet.column_dimensions[get_column_letter(column_index)].width = column_width + 5

        # --- Freeze
        if freeze_cell:
            cell_name = get_column_letter(freeze_cell[1]) + str(freeze_cell[0])
            LOGGER.info(f'Freezing cell "{cell_name}" ...')
            sheet.freeze_panes = cell_name

        # Save workbook
        LOGGER.info(f'Saving workbook to "{filename}" ...')
        workbook.save(filename)

        return


def main():
    departments = {}
    workers = {}
    tasks = {}

    # Read excel workbook
    workbook = openpyxl.load_workbook(INPUT_FILENAME, read_only=True)
    sheet = workbook.active
    sheet_values = list(sheet.values)
    workbook.close()
    raw_book_data = [{sheet_values[0][i]: row[i] for i in range(len(sheet_values[0]))} for row in sheet_values[1:]]
    group = None
    worker_username = None

    for row in raw_book_data:
        key = row['Key']
        task_name = row['Summary']
        task_status = row['Status']
        assignee = row['Assignee']
        start_date = row['Start Date [Gantt]']
        end_date = row['End Date [Gantt]']
        estimate = row['Original Estimate']
        priority = row['Priority']

        if key is None:
            group = task_name
            if group not in departments:
                department = Department(group)
                departments[group] = department
            continue

        if task_status is None:
            worker_username = key

        # Skip all rows where "key" is not actual Jira task key
        if (key is None) or (not Task.is_jira_key(key)):
            continue

        # Skip all tasks without start_date and end_date
        if (not start_date) or (not end_date):
            continue

        # Skip all tasks where start_date or end_date not in requested year
        if (start_date.year != YEAR) or (end_date.year != YEAR):
            # assert start_date.year == YEAR, f"{start_date.year} != {YEAR}"
            # assert end_date.year == YEAR, f"{end_date.year} != {YEAR}"
            continue

        # Skip all tasks without estimate
        if not estimate:
            continue

        assignee = assignee.strip()
        task_name = task_name.strip()
        estimate = estimate.days * 24 + estimate.seconds // 60 // 60

        # Create new task
        if key not in tasks:
            task = Task(
                name=task_name,
                key=key,
                date_start=start_date.date(),
                date_end=end_date.date(),
                priority=priority,
                status=task_status,
                estimate=estimate,
            )
            tasks[key] = task

        # Create new worker
        if assignee not in workers:
            worker = Worker(
                assignee,
                username=worker_username,
            )
            workers[assignee] = worker

        # Add task to worker
        workers[assignee].add_task(
            task=tasks[key],
        )

        # Add worker to department
        if workers[assignee] not in departments[group].workers:
            departments[group].add_worker(workers[assignee])

    departments = [department for department in departments.values() if department.workers]
    workers = [worker for worker in workers.values() if worker.tasks]
    scheduler = Scheduler(
        departments=departments,
        workers=workers,
        tasks=list(tasks.values()),
    )
    generated_data, merge_data, style_data = scheduler.generate_data(YEAR)
    scheduler.write_data_to_excel(
        data=generated_data,
        merge_data=merge_data,
        style_data=style_data,
        freeze_cell=FREEZE_CELL,
        filename=OUTPUT_FILENAME,
    )
    print(f'Data successfully generated to "{OUTPUT_FILENAME}"!')


class CheckArgs:
    @staticmethod
    def input_filename(value: str) -> str:
        """
        Check path existing and file type

        :param value: Any path/filename

        :return: Given path/filename
        """
        if not os.path.exists(value):
            raise ArgumentTypeError(f'File is not exists: "{value}"')

        if not value.endswith('.xlsx'):
            raise ArgumentTypeError(f'Input filename must be xlsx format')

        return value

    @staticmethod
    def output_filename(value: str) -> str:
        """
        Check output file type

        :param value: Any path/filename

        :return: Given path/filename
        """
        if not value.endswith('.xlsx'):
            raise ArgumentTypeError(f'Output filename must be xlsx format')

        return value

    @staticmethod
    def year(value: str) -> int:
        """
        Check year value

        :param value: Year

        :return: Year
        """
        if not all([value.isnumeric(), value.isdecimal(), value.isdigit()]):
            raise ArgumentTypeError('Year must be a number')
        value = int(value)

        return value

    @staticmethod
    def freeze_cell(value: str) -> _T_CORDS:
        """
        Check freeze cell value

        :param value: Freeze cell cords str

        :return: Freeze cell cords tuple
        """
        error_msg = 'Cell cords must be two numbers divided by a comma (ex: "1,1" or "12, 34")'

        if ',' not in value:
            raise ArgumentTypeError(error_msg)
        value = value.split(',')
        if len(value) != 2:
            raise ArgumentTypeError(error_msg)
        try:
            value = (int(value[0].strip()), int(value[1].strip()))
        except ValueError:
            raise ArgumentTypeError(error_msg)

        return value


if __name__ == '__main__':
    # Logging
    LOGGING_FORMAT = '[%(asctime)s] %(levelname)s: %(filename)s(%(lineno)d): %(funcName)s: %(message)s'
    logging.basicConfig(
        filename=__file__[:__file__.rfind('.')] + '.log',
        filemode='w',
        format=LOGGING_FORMAT,
        level=logging.INFO,
        encoding='utf-8',
    )
    LOGGER = logging.getLogger(__name__)

    # Parse args
    ArgParser = ArgumentParser()
    ArgParser.add_argument(
        'filename',
        help='Input filename (must be xlsx)',
        type=CheckArgs.input_filename,
    )
    ArgParser.add_argument(
        '-o', '--output-filename',
        default='output.xlsx',
        help='Output filename (must be xlsx)',
        type=CheckArgs.output_filename,
    )
    ArgParser.add_argument(
        '-y', '--year',
        default=datetime.today().year,
        help='Year for data generation',
        type=CheckArgs.year,
    )
    ArgParser.add_argument(
        '--freeze-cell',
        default=(4, 4),
        help='Cell cords for freezing rows above and columns to the left (ex: "3,3" or "12, 34")',
    )
    ScriptArgs = ArgParser.parse_args()
    INPUT_FILENAME = ScriptArgs.filename
    OUTPUT_FILENAME = ScriptArgs.output_filename
    YEAR = ScriptArgs.year
    FREEZE_CELL = ScriptArgs.freeze_cell

    try:
        main()
    except BaseException as main_exception:
        LOGGER.exception('An exception has occurred:')
        LOGGER.exception(main_exception)
        raise
